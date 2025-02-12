import os
import configparser
import shutil
from datetime import datetime, timedelta
import openpyxl
from copy import copy
from openpyxl.utils.exceptions import InvalidFileException
from openpyxl.utils import get_column_letter, column_index_from_string
import logging
import subprocess

# ===========================
# = Manejo del archivo conf =
# ===========================
def modificar_valor_conf(ruta_archivo, clave, nuevo_valor):
    with open(ruta_archivo, 'r') as archivo:
        lineas = archivo.readlines()

    with open(ruta_archivo, 'w') as archivo:
        for linea in lineas:
            if linea.startswith(clave + " ="):
                archivo.write(f"{clave} = {nuevo_valor}\n")
            else:
                archivo.write(linea)

def limpiar_archivo(ruta_archivo):
    with open(ruta_archivo, 'r', encoding='ascii') as archivo:
        contenido = archivo.read().replace('\n', '\r\n')
    contenido_limpio = contenido.rstrip()
    with open(ruta_archivo, 'w', encoding='ascii') as archivo:
        archivo.write(contenido_limpio)

def leer_conf(ruta_completa):
    limpiar_archivo(ruta_completa)
    config = configparser.ConfigParser()
    
    try:
        archivos_leidos = config.read(ruta_completa)

        if not archivos_leidos:
            raise FileNotFoundError(f"El archivo {ruta_completa} no se pudo leer.")
        
        if 'Marcado' not in config:
            raise configparser.NoSectionError('Marcado')
        
        if ruta_completa.endswith('.conf'):
            asunto = config.get('Marcado', 'asunto')
            asunto_lista = asunto.split(',')
            fecha = config.get('Marcado', 'fecha')
            fecha_dt = datetime.strptime(fecha, '%d/%m/%Y').date()
            return asunto_lista, fecha_dt

        elif ruta_completa.endswith('.lobo'):
            marcado = config.get('Marcado', 'marcado')
            marcado_lista = marcado.split(",")
            return marcado_lista, None
        
    except configparser.NoSectionError as e:
        print(f"Error: La sección no existe en el archivo de configuración: {e}")
        return None, None  
    except configparser.NoOptionError as e:
        print(f"Error: La opción no existe en la sección especificada: {e}")
        return None, None  
    except FileNotFoundError as e:
        print(f"Error al leer el archivo: {e}")
        return None, None  
    except Exception as e:
        print(f"Error inesperado en leer_conf: {e}")
        return None, None 

# ===========================
# = Manejo del archivo jar =
# ===========================

def ejecucionJava(ruta_actual, ruta_java):
    """
    Ejecuta un archivo JAR de Java pasando parámetros específicos y maneja la salida y errores del proceso.

    :param ruta_actual: Ruta del directorio actual donde se encuentra el archivo de configuración.
    :param ruta_java: Ruta completa del archivo JAR de Java a ejecutar.
    """
    try:
        cmd = ['java', '-jar', ruta_java ] 
        execution_process = subprocess.Popen(cmd, stdout=subprocess.PIPE, stderr=subprocess.PIPE)

        execution_output, execution_errors = execution_process.communicate()
        
        if execution_errors:
            exit_code = execution_process.returncode
            cmd_executed = " ".join(cmd)
            stdout_output = execution_output.decode('utf-8')
            logging.error(f"Error de ejecución Java con el comando '{cmd_executed}'. \nCódigo de salida: {exit_code}. \nError: {execution_errors.decode('utf-8')}. \nSalida estándar: {stdout_output}")
            exit(1)

        logging.info(execution_output.decode('utf-8'))
        
    except AttributeError as ae:
        logging.error(f"Error de atributo: {ae}")
    except subprocess.CalledProcessError as cpe:
        logging.error(f"Error al llamar al proceso Java: {cpe.output.decode('utf-8')}")
    except Exception as e:
        logging.error(f"Error de ejecución: {e}")

def cargarAltex(ruta_actual):
    """
    Carga la configuración desde un archivo específico y retorna los valores necesarios para la ejecución de Java.

    :param ruta_actual: Ruta del directorio actual donde se encuentra el archivo de configuración.
    :return: Una tupla que contiene la fecha y los marcados obtenidos del archivo de configuración.
    """
    configAltex_ruta = os.path.join(ruta_actual, 'Marcado_altex.lobo')
    limpiar_archivo(configAltex_ruta)
    try:
        configAltex = configparser.ConfigParser()
        if not configAltex.read(configAltex_ruta):
            raise FileNotFoundError(f"No se pudo encontrar o leer el archivo de configuración: {configAltex_ruta}")
        
        fecha_altex = configAltex.get('Marcado', 'fecha')
        marcado_altex = configAltex.get('Marcado', 'marcado').split(",")

        return fecha_altex, marcado_altex
    
    except FileNotFoundError as fnf_error:
        print(f"Archivo de configuración no encontrado: {fnf_error}")
    except configparser.NoSectionError as nse_error:
        print(f"Sección 'Marcado' no encontrada en el archivo de configuración: {nse_error}")
    except configparser.NoOptionError as noe_error:
        print(f"Una opción requerida está ausente en la sección 'Marcado': {noe_error}")
    except Exception as e:
        print(f"Error al cargar configuración Altex: {e}")

# ===========================
# = Manejo de Filas Excel =
# ===========================

def insertar_fila(sheet, now):
    """
    Inserta una nueva fila en la posición 8 de la hoja de Excel y copia el estilo de la fila siguiente.
    Además, establece la fecha actual en la primera celda de la nueva fila.

    :param sheet: La hoja de Excel donde se insertará la fila.
    :param now: Objeto datetime con la fecha actual.
    """
    try:
        sheet.insert_rows(8)
        for col in range(1, sheet.max_column + 1):
            celda_vieja = sheet.cell(row=9, column=col)
            celda_nueva = sheet.cell(row=8, column=col)

            celda_nueva.font = copy(celda_vieja.font)
            celda_nueva.fill = copy(celda_vieja.fill)
            celda_nueva.alignment = copy(celda_vieja.alignment)

        fecha_formateada = datetime.strftime(now, "%d/%m/%Y")
        sheet.cell(row=8, column=1).value = fecha_formateada

    except InvalidFileException as e:
        logging.error(f"Error al trabajar con el archivo de Excel: {e}")
    except Exception as e:
        logging.error(f"Ocurrió un error inesperado: {e}")

def copiar_fila(origen_row, origen_col, destino_row, destino_col, sh):
    """
    :param origen_row: La fila de la celda origen.
    :param origen_col: La columna de la celda origen.
    :param destino_row: La fila de la celda destino.
    :param destino_col: La columna de la celda destino.
    :param sh: La hoja de Excel donde se encuentran las celdas.
    """
    celda_origen = sh.cell(row=origen_row, column=origen_col)
    celda_destino = sh.cell(row=destino_row, column=destino_col)

    celda_destino.font = copy(celda_origen.font)
    celda_destino.fill = copy(celda_origen.fill)
    celda_destino.alignment = copy(celda_origen.alignment)

def crear_archivo_excel_si_no_existe(carpeta_excel, ruta_excel_base, now):
    """
    :param carpeta_excel: Ruta de la carpeta donde se creará el archivo Excel.
    :param ruta_excel_base: Ruta del archivo Excel base que se copiará.
    :param now: Objeto datetime que representa el momento actual.
    :return: Ruta del archivo Excel creado o existente.
    """
    try:
        archivo_excel_year = os.path.join(carpeta_excel, f"Bitacora_de_respaldos_BD_{now.year}.xlsx")

        if not os.path.exists(archivo_excel_year):
            logging.info(f"Creando archivo excel en la ruta: {carpeta_excel}")
            shutil.copy2(ruta_excel_base, archivo_excel_year)
        else:
            logging.info(f"El archivo excel ya existe en la ruta: {archivo_excel_year}")
    except (FileNotFoundError, PermissionError) as e:
        logging.error(f"Error al crear o verificar el archivo Excel: {e}")
        return None
    return archivo_excel_year

# ================================
# = Llenado de bitacora en excel =
# ================================

def comprobar_backup(now, esquemas, ruta_base, letras, sheet, marcado_lista):
    """
    Comprueba la existencia de backups para una lista de esquemas y marca en una hoja de Excel si se encuentran.

    :param now: Objeto datetime con la fecha actual.
    :param esquemas: Lista de nombres de esquemas a comprobar.
    :param ruta_base: Ruta base donde se encuentran los folders de los esquemas.
    :param letras: Lista de letras de columnas en la hoja de Excel para marcar.
    :param sheet: Objeto de hoja de Excel donde se realizarán las marcas.
    """
    try:
        ruta_folder_mg = os.path.join(ruta_base, 'Mongo')
        
        for i, esquema in enumerate(esquemas):
            ruta_folder = os.path.join(ruta_base, esquema)

            if esquema in {'SIAL_HDE', 'SialCFDI', 'CAMPOBDB'} and os.path.exists(ruta_folder_mg):
                marcado_de_esquemas(ruta_folder_mg, i, esquema, now, sheet, letras, None)
            elif os.path.exists(ruta_folder):
                marcado_de_esquemas(ruta_folder, i, esquema, now, sheet, letras, None)
            else:
                marcado_de_esquemas(ruta_folder, i, esquema, now, sheet, letras, marcado_lista)

    except FileNotFoundError as e:
        logging.error(f"Archivo o directorio no encontrado: {e}")
    except Exception as e:
        logging.error(f"Error inesperado al comprobar backups: {e}")

def marcado_de_esquemas(ruta_folder, idx, esquema, now, sheet, letras, marcado_lista):
    try:
        lista_mongo = ['SIAL_HDE','SialCFDI','CAMPOBDB']
        # cambiar para leer directo

        asunto_cadena = 'SIAL_ALTEX ,SIAL_ALTEX_FREX ,SIAL_ALTEX_ALXTRA ,SIAL_ALTEX_NEXT ,SIAL_ALTEX_XTRA ,SIALADMIN_ALTEX ,SIALADMIN_ALTEX_ALXTRA ,SIALADMIN_ALTEX_FREX ,SIALADMIN_ALTEX_NEXT ,SIALADMIN_ALTEX_XTRA '
        asunto = asunto_cadena.split(',')
        if esquema == 'SEGUIDORES' or esquema == 'REPCIU_AYTO_ZAMORA':
            archivo = f"{esquema}-respaldo-{now.strftime('%Y%m%d')}.tar.gz"            
        elif esquema == 'ZAM-SV-MORPHO2':
            archivo = f"MorphoManager-{now.strftime('%Y%m%d')}.7z"
        elif esquema in lista_mongo:
            archivo = f"{now.strftime('%Y%m%d')}-{esquema}.7z"
        else:
                archivo = f"{esquema}-respaldo-{now.strftime('%Y%m%d')}.tar.gz"

        ruta_archivo = os.path.join(ruta_folder, archivo)

        if os.path.exists(ruta_archivo):
            sheet.cell(row=8, column=column_index_from_string(letras[idx])).value = 'X'
            # print(f"El backup del esquema {esquema} se realizo correctamente")

        for j, value in enumerate(asunto):
            if esquema in asunto[j]: # poner otro condicional
                if 'Si' in marcado_lista[j]:
                    # poner un break despues de marcar el esquema, no tiene sentido seguir iterando si para ese esquema ya se marco
                    sheet.cell(row=8, column=column_index_from_string(letras[idx])).value = 'X'

    except FileNotFoundError as e:
        logging.error(f"Archivo de backup no encontrado: {e}")
    except Exception as e:
        logging.error(f"Error inesperado al marcar esquemas: {e}")

# ===========================
# = Manejo del archivo .ini =
# ===========================

# def escribir_configuracion(ruta_archivo, esquemas, letras):
#     config = configparser.ConfigParser()
#     config['ESQUEMAS'] = {
#         'esquema': ','.join(esquemas),
#         'letras': ','.join(letras)
#     }
    
#     with open(ruta_archivo, 'w') as configfile:
#         config.write(configfile)

def leer_configuracion(ruta_archivo):
    config = configparser.ConfigParser()
    config.read(ruta_archivo)
    
    esquemas = config['ESQUEMAS']['esquema'].split(',')
    letras = config['ESQUEMAS']['letras'].split(',')
    
    return esquemas, letras

def actualizar_esquemas_y_letras(ruta_base, esquemas, letras, lista_mongo, asunto):
    esquema_set = set(esquemas) 
    pos_final_num = column_index_from_string(letras[-1]) 

    nombres_nuevos, letras_nuevas = [], []
    ctr = 1

    exclusiones = {'SIAL_PRUEBA', 'LOBORH', 'FREXPORT', 'Mongo'}

    inclusiones = asunto

    try:
        dirs_to_process = [d for d in os.listdir(ruta_base) 
                           if os.path.isdir(os.path.join(ruta_base, d)) and 
                           d not in esquema_set and 
                           d not in exclusiones]
        dirs_to_process.extend(subdir for subdir in lista_mongo if subdir not in esquema_set)
        dirs_to_process.extend(inc for inc in inclusiones if inc not in esquema_set)

        for nombre_dir in dirs_to_process:
            nombres_nuevos.append(nombre_dir)
            letras_nuevas.append(get_column_letter(pos_final_num + ctr))
            ctr += 1

    except PermissionError as e:
        print(f"No se pudo acceder a la ruta {ruta_base}: {e}")
        
    return nombres_nuevos, letras_nuevas

def actualizar_archivo_de_config(ruta_base, esquemas, letras, lista_mongo, asunto):
    nombres_nuevos, letras_nuevas = actualizar_esquemas_y_letras(ruta_base, esquemas, letras, lista_mongo, asunto)

    if nombres_nuevos or letras_nuevas:
        esquemas.extend(nombres_nuevos)
        letras.extend(letras_nuevas)

        config = configparser.ConfigParser()
        path = os.path.join(ruta_base, 'config.ini')
        config.read(path)

        esquema_str = ','.join(esquemas)
        letra_str = ','.join(letras)

        config.set('ESQUEMAS', 'esquema', esquema_str)
        config.set('ESQUEMAS', 'letras', letra_str)

        with open(path, 'w', encoding='ascii') as configfile:
            config.write(configfile)

        print(f"Se añadieron los siguientes esquemas: {nombres_nuevos}")
        print(f"Se añadieron las siguientes letras: {letras_nuevas}")

    return esquemas, letras

# ===========================
# = Llamado a las funciones =
# ===========================

def main_function(sheet, now, esquemas, ruta_base, letras, marcado_lista):
    insertar_fila(sheet, now)

    comprobar_backup(now, esquemas, ruta_base, letras, sheet, marcado_lista)
    
    # se añaden los nombres de las carpetas a la bitácora
    # esto deberia ir primero y despues checar los backups
    for i, letra in enumerate(letras):
        if esquemas[i] not in ['Mongo', 'LOBORH', 'FREXPORT', 'SIAL_PRUEBA']:            
            copiar_fila(7, column_index_from_string(letra) - 1, 7, column_index_from_string(letra), sheet)
            sheet.cell(row=7, column=column_index_from_string(letra)).value = esquemas[i]

# =================
# = Logica inical =
# =================

# ruta actual: /home/julianq@local.lobos.com.mx/Documents/Python/Bitacora_BD_v2
ruta_actual = os.path.dirname(__file__)
nombre_excel_base = "Bitacora_de_respaldos_BD"
carpeta_excel = os.path.join(ruta_actual, nombre_excel_base)
ruta_excel_base = os.path.join(ruta_actual, nombre_excel_base + ".xlsx")

ruta_java = os.path.join(ruta_actual, "prueba", 'marcarAltex.jar')

lista_mongo = {'SIAL_HDE','SialCFDI','CAMPOBDB'}

file_path_conf = os.path.join(ruta_actual, "configMarcado.conf")
file_path_lobo = os.path.join(ruta_actual, "Marcado_altex.lobo")
ruta_archivo_config = os.path.join(ruta_actual, "config.ini")

asunto, fecha_modificacion = leer_conf(file_path_conf)

ruta_base = '/mnt/175/'
#ruta_base = r'\\192.0.0.175\respaldos'

esquemas, letras = leer_configuracion(ruta_archivo_config)
esquemas, letras = actualizar_archivo_de_config(ruta_base, esquemas, letras, lista_mongo, asunto)

if len(letras) != len(esquemas):
    raise ValueError("La cantidad de letras y esquemas no coincide")

now = datetime.now().date()
# fecha_modificacion = datetime.strptime(fecha, '%d/%m/%Y').date()  # Fecha que se aumenta en cada iteracion
# now = datetime(2024, 7, 21).date() # Fecha de prueba

os.makedirs(carpeta_excel, exist_ok=True)

archivo_excel_year = crear_archivo_excel_si_no_existe(carpeta_excel, ruta_excel_base, fecha_modificacion)

workbook = openpyxl.load_workbook(archivo_excel_year)
sheet = workbook.active

llenar_bitacora = True # Variable para controlar si se debe llenar la bitácora o no

# marcado_lista, fecha = leer_conf(file_path_lobo)
# la fecha dada es de cuando debe de marcar

if llenar_bitacora:
    while fecha_modificacion < now:
        ejecucionJava(ruta_actual, ruta_java)
        marcado_lista, _ = leer_conf(file_path_lobo)
        main_function(sheet, fecha_modificacion, esquemas, ruta_base, letras, marcado_lista)

        fecha_modificacion += timedelta(days=1)
        fecha_mod_str = fecha_modificacion.strftime('%d/%m/%Y')
        modificar_valor_conf(file_path_conf, 'fecha', fecha_mod_str)
        
        
        limpiar_archivo(file_path_conf)
                
else:
    limpiar_archivo(file_path_conf)
    ejecucionJava(ruta_actual, ruta_java)
    marcado_lista, _ = leer_conf(file_path_lobo)
    main_function(sheet, now, esquemas, ruta_base, letras, marcado_lista)

workbook.save(archivo_excel_year)
